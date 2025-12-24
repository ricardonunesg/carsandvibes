from openpyxl import load_workbook
from pathlib import Path

SRC = Path.home() / "carsandvibes/imports/source/2026_RRP_OMP_23102025_variants_options_prepared.xlsx"
OUT = SRC.with_name(SRC.stem + "_NO_FACETS.xlsx")

# Headers que normalmente aparecem em imports (ajusta se precisares)
MATCH_SUBSTRINGS = [
    "facet", "facetvalue", "facet value", "facet_value",
    "facetvalueids", "facet_value_ids", "facetValueIds",
    "facets", "facetvalues", "facet values",
]

def header_matches(value: str) -> bool:
    if value is None:
        return False
    s = str(value).strip().lower()
    return any(m in s for m in MATCH_SUBSTRINGS)

wb = load_workbook(SRC)
total_cleared_cells = 0
total_cleared_cols = 0

for ws in wb.worksheets:
    # assume 1ª linha = headers
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=1, column=col).value)

    facet_cols = [i+1 for i, h in enumerate(headers) if header_matches(h)]
    if not facet_cols:
        continue

    total_cleared_cols += len(facet_cols)
    for col in facet_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                total_cleared_cells += 1
                cell.value = None

    print(f"[{ws.title}] colunas limpas: {len(facet_cols)} -> {', '.join(str(c) for c in facet_cols)}")

wb.save(OUT)
print("\nOK!")
print(f"Ficheiro novo: {OUT}")
print(f"Colunas limpas (total): {total_cleared_cols}")
print(f"Células limpas (total): {total_cleared_cells}")
